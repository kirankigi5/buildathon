
import io
import re
import csv
from typing import List, Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from schemas import StartupInput, StartupResult

# Keyword mapping for fuzzy column detection
FIELD_KEYWORDS = {
	"name": ["company name", "company", "startup name", "startup", "name", "companies", "organization"],
	"description": ["description", "desc", "about", "summary", "overview", "pitch", "business description"],
	"industry": ["industry", "sector", "primary industry", "vertical", "category", "primary industry group"],
	"stage": ["stage", "financing stage", "deal type", "round", "last financing deal type", "series"],
	"founder_name": ["founder", "co-founder", "founders", "primary contact", "ceo", "contact name"],
	"linkedin_url": ["linkedin", "linkedin url", "linkedin profile"],
	"website": ["website", "url", "web", "site", "homepage", "company url"],
	"total_raised_m": ["total raised", "amount raised", "raised", "total funding", "capital raised", "funding amount"],
	"location": ["location", "city", "hq", "headquarters", "hq location", "hq city"],
	"employees": ["employees", "headcount", "team size", "employee count"],
	"traction": ["traction", "revenue", "mrr", "arr", "growth", "metrics"],
	"competitors": ["competitors", "competition", "alternatives"],
}

def normalize_col(col):
	return re.sub(r'[^a-z0-9 ]', '', col.lower().strip())

def detect_header_row(rows: List[List[str]]) -> int:
	# Scan first 10 rows for header keywords
	best_row = 0
	best_score = 0
	for i, row in enumerate(rows[:10]):
		score = 0
		for cell in row:
			cell_norm = normalize_col(cell)
			for keywords in FIELD_KEYWORDS.values():
				for kw in keywords:
					kw_norm = normalize_col(kw)
					if kw_norm in cell_norm:
						score += 1
		if score > best_score:
			best_score = score
			best_row = i
	return best_row

def detect_column_mapping(header: List[str]) -> Dict[str, str]:
	mapping = {}
	header_norm = [normalize_col(h) for h in header]
	for field, keywords in FIELD_KEYWORDS.items():
		found = None
		for kw in keywords:
			kw_norm = normalize_col(kw)
			# Priority 1: exact match
			for i, h in enumerate(header_norm):
				if h == kw_norm:
					found = header[i]
					break
			if found:
				break
		if not found:
			# Priority 2: startswith
			for i, h in enumerate(header_norm):
				if h.startswith(kw_norm):
					found = header[i]
					break
			if found:
				break
		if not found:
			# Priority 3: contains
			for i, h in enumerate(header_norm):
				if kw_norm in h:
					found = header[i]
					break
			if found:
				break
		mapping[field] = found
	return mapping

def clean_total_raised(val):
	if not val:
		return 0.0
	val = str(val).replace("$", "").replace(",", "").replace("M", "").replace("m", "")
	try:
		return float(val)
	except Exception:
		return 0.0

def parse_excel(file_content: bytes) -> Tuple[List[StartupInput], Dict[str, str]]:
	# Detect file type
	if file_content[:4] == b'PK\x03\x04':  # XLSX magic bytes
		df = pd.read_excel(io.BytesIO(file_content), header=None)
		rows = df.values.tolist()
	else:
		# Assume CSV
		text = file_content.decode('utf-8', errors='replace')
		reader = csv.reader(io.StringIO(text))
		rows = list(reader)
	header_idx = detect_header_row(rows)
	header = rows[header_idx]
	mapping = detect_column_mapping(header)
	startups = []
	for row in rows[header_idx+1:]:
		if len(row) < len(header):
			row += [""] * (len(header) - len(row))
		data = dict(zip(header, row))
		name = data.get(mapping["name"] or "", "").strip()
		description = data.get(mapping["description"] or "", "").strip()
		if not name and not description:
			continue
		startup = StartupInput(
			name=name or "Unknown",
			description=description or "No description",
			industry=data.get(mapping["industry"] or "", "Unknown"),
			stage=data.get(mapping["stage"] or "", "Seed"),
			founder_name=data.get(mapping["founder_name"] or "", "Not provided"),
			linkedin_url=data.get(mapping["linkedin_url"] or "", ""),
			website=data.get(mapping["website"] or "", ""),
			total_raised_m=clean_total_raised(data.get(mapping["total_raised_m"] or "", 0.0)),
			location=data.get(mapping["location"] or "", ""),
			employees=int(data.get(mapping["employees"] or "0") or 0),
			traction=data.get(mapping["traction"] or "", "Not disclosed"),
			competitors=data.get(mapping["competitors"] or "", "")
		)
		startups.append(startup)
		if len(startups) >= 10:
			break
	return startups, mapping

def create_output_excel(results: List[StartupResult]) -> bytes:
	wb = Workbook()
	ws = wb.active
	ws.title = "TierVC Results"
	header = [
		"Company Name", "Tier", "Score (0-100)", "Invest?", "Confidence",
		"Top Pro Argument", "Top Risk", "Founder", "LinkedIn", "Website"
	]
	ws.append(header)
	# Bold header
	for cell in ws[1]:
		cell.font = Font(bold=True)
	# Sort by score descending
	results_sorted = sorted(results, key=lambda r: r.score, reverse=True)
	tier_colors = {
		1: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Green
		2: PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # Yellow
		3: PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),  # Red
	}
	for r in results_sorted:
		row = [
			r.name, r.tier, r.score, r.invest, r.confidence,
			r.top_pro, r.top_risk, r.founder_name, r.linkedin_url, r.website
		]
		ws.append(row)
		tier = r.tier
		fill = tier_colors.get(tier, None)
		if fill:
			for cell in ws[ws.max_row]:
				cell.fill = fill
	# Set column widths
	col_widths = [20, 8, 12, 10, 10, 30, 30, 15, 30, 30]
	for i, width in enumerate(col_widths, 1):
		ws.column_dimensions[get_column_letter(i)].width = width
	output = io.BytesIO()
	wb.save(output)
	return output.getvalue()
